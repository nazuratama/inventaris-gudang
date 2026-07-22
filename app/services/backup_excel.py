"""Excel workbook creation, SQLite snapshots, and backup retention."""

from __future__ import annotations

import logging
import os
import shutil
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app import __version__
from app.errors import AppError
from app.infrastructure.database import Database
from app.utils import (
    MONEY_SCALE,
    excel_safe_text,
    new_id,
    raw_to_quantity,
    utc_now,
)

logger = logging.getLogger("inventory.backup")

ITEM_HEADERS = [
    "Item ID",
    "Name",
    "Category",
    "Location",
    "Unit",
    "Current Stock",
    "Description",
    "Created Time",
    "Updated Time",
]
# Older workbooks still accepted on import (Active Status ignored; always active).
LEGACY_ITEM_HEADERS = [
    "Item ID",
    "SKU",
    "Name",
    "Category",
    "Location",
    "Unit",
    "Current Stock",
    "Minimum Stock",
    "Stock Status",
    "Purchase Price",
    "Selling Price",
    "Currency Code",
    "Description",
    "Active Status",
    "Data Type",
    "Created Time",
    "Updated Time",
]
LEGACY_ITEM_HEADERS_V1 = [
    "Item ID",
    "SKU",
    "Name",
    "Category",
    "Location",
    "Unit",
    "Current Stock",
    "Minimum Stock",
    "Stock Status",
    "Description",
    "Active Status",
    "Created Time",
    "Updated Time",
]
LEGACY_ITEM_HEADERS_V2 = [
    "Item ID",
    "Name",
    "Category",
    "Location",
    "Unit",
    "Current Stock",
    "Description",
    "Active Status",
    "Data Type",
    "Created Time",
    "Updated Time",
]
# Previous simple export that still included Active Status.
LEGACY_ITEM_HEADERS_V3 = [
    "Item ID",
    "Name",
    "Category",
    "Location",
    "Unit",
    "Current Stock",
    "Description",
    "Active Status",
    "Created Time",
    "Updated Time",
]
MOVEMENT_HEADERS = [
    "Movement ID",
    "Item ID",
    "Date",
    "Item Name",
    "Type",
    "Quantity",
    "Stock Before",
    "Stock After",
    "Note",
]
LEGACY_MOVEMENT_HEADERS = [
    "Movement ID",
    "Item ID",
    "Date",
    "SKU",
    "Item Name",
    "Type",
    "Quantity",
    "Stock Before",
    "Stock After",
    "Note",
    "Reference Number",
    "Data Type",
]
LEGACY_MOVEMENT_HEADERS_V1 = [
    "Movement ID",
    "Item ID",
    "Date",
    "SKU",
    "Item Name",
    "Type",
    "Quantity",
    "Stock Before",
    "Stock After",
    "Note",
    "Reference Number",
]
LEGACY_MOVEMENT_HEADERS_V2 = [
    "Movement ID",
    "Item ID",
    "Date",
    "Item Name",
    "Type",
    "Quantity",
    "Stock Before",
    "Stock After",
    "Note",
    "Data Type",
]
CATEGORY_HEADERS = ["Category ID", "Name", "Created Time", "Updated Time"]
LEGACY_CATEGORY_HEADERS = ["Category ID", "Name", "Data Type", "Created Time", "Updated Time"]
LOCATION_HEADERS = [
    "Location ID",
    "Name",
    "Description",
    "Created Time",
    "Updated Time",
]
LEGACY_LOCATION_HEADERS = [
    "Location ID",
    "Name",
    "Description",
    "Data Type",
    "Created Time",
    "Updated Time",
]
class ExcelBackupService:
    """Generate a verified workbook without writing over the last valid file directly."""

    def __init__(self, database: Database) -> None:
        self.database = database
        self.config = database.config

    def create(self) -> dict[str, Any]:
        timestamp = datetime.now(UTC)
        stamp = timestamp.strftime("%Y-%m-%d_%H-%M-%S")
        current_path = self.config.current_excel_backup_path
        daily_path = self.config.daily_backups_path / f"inventory_{stamp}.xlsx"
        temporary_path = current_path.with_name(f".{current_path.name}.{new_id()}.tmp.xlsx")
        daily_temporary = daily_path.with_suffix(f".{new_id()}.tmp.xlsx")
        self.config.backups_path.mkdir(parents=True, exist_ok=True)
        self.config.daily_backups_path.mkdir(parents=True, exist_ok=True)
        log_id = self._start_log(current_path)
        try:
            workbook = self._build_workbook(timestamp)
            workbook.save(temporary_path)
            workbook.close()
            self._verify_workbook(temporary_path)

            shutil.copy2(temporary_path, daily_temporary)
            self._verify_workbook(daily_temporary)
            os.replace(daily_temporary, daily_path)
            os.replace(temporary_path, current_path)
            self._finish_log(log_id, "SUCCESS", None)
            self._cleanup_retention()
            logger.info("Excel backup created: %s", daily_path.name)
            return {
                "status": "SUCCESS",
                "file_name": current_path.name,
                "daily_file_name": daily_path.name,
                "created_at": utc_now(),
            }
        except Exception as exc:
            for path in (temporary_path, daily_temporary):
                try:
                    path.unlink(missing_ok=True)
                except OSError:
                    logger.warning("Could not remove failed backup temporary file %s", path.name)
            technical = f"{type(exc).__name__}: {str(exc)[:300]}"
            self._finish_log(log_id, "FAILED", technical)
            logger.exception("Excel backup failed; the previous current backup was preserved")
            raise AppError(
                "BACKUP_FAILED",
                "Backup Excel gagal. Data utama tetap tersimpan di database.",
                status_code=500,
            ) from exc

    def _query_data(self) -> dict[str, Any]:
        with self.database.connection() as connection:
            connection.execute("BEGIN")
            try:
                items = connection.execute("""
                    SELECT i.*, c.name AS category_name, l.name AS location_name
                    FROM items i
                    LEFT JOIN categories c ON c.id = i.category_id
                    LEFT JOIN locations l ON l.id = i.location_id
                    ORDER BY i.name COLLATE NOCASE
                    """).fetchall()
                movements = connection.execute("""
                    SELECT m.*, i.name AS item_name
                    FROM stock_movements m JOIN items i ON i.id = m.item_id
                    ORDER BY m.created_at, m.rowid
                    """).fetchall()
                categories = connection.execute("""
                    SELECT COALESCE(c.name, 'Tanpa kategori') AS name,
                           COUNT(i.id) AS item_count,
                           COALESCE(SUM(i.current_stock), 0) AS total_stock
                    FROM items i LEFT JOIN categories c ON c.id = i.category_id
                    GROUP BY c.id ORDER BY name
                    """).fetchall()
                locations = connection.execute("""
                    SELECT COALESCE(l.name, 'Tanpa lokasi') AS name,
                           COUNT(i.id) AS item_count,
                           COALESCE(SUM(i.current_stock), 0) AS total_stock
                    FROM items i LEFT JOIN locations l ON l.id = i.location_id
                    GROUP BY l.id ORDER BY name
                    """).fetchall()
                category_records = connection.execute(
                    "SELECT * FROM categories ORDER BY name COLLATE NOCASE"
                ).fetchall()
                location_records = connection.execute(
                    "SELECT * FROM locations ORDER BY name COLLATE NOCASE"
                ).fetchall()
                settings = {
                    row["key"]: row["value"]
                    for row in connection.execute("SELECT key, value FROM app_settings").fetchall()
                }
                connection.commit()
            except Exception:
                connection.rollback()
                raise
        return {
            "items": items,
            "movements": movements,
            "categories": categories,
            "locations": locations,
            "category_records": category_records,
            "location_records": location_records,
            "settings": settings,
        }

    def _build_workbook(self, created_at: datetime) -> Workbook:
        data = self._query_data()
        workbook = Workbook()
        items_sheet = workbook.active
        items_sheet.title = "Items"
        movements_sheet = workbook.create_sheet("Stock Movements")
        categories_sheet = workbook.create_sheet("Categories")
        locations_sheet = workbook.create_sheet("Locations")
        summary_sheet = workbook.create_sheet("Summary")
        metadata_sheet = workbook.create_sheet("Backup Metadata")
        settings_sheet = workbook.create_sheet("Advanced Settings")

        items_sheet.append(ITEM_HEADERS)
        for row in data["items"]:
            current = int(row["current_stock"])
            items_sheet.append(
                [
                    row["id"],
                    excel_safe_text(row["name"]),
                    excel_safe_text(row["category_name"]),
                    excel_safe_text(row["location_name"]),
                    excel_safe_text(row["unit"]),
                    raw_to_quantity(current),
                    excel_safe_text(row["description"]),
                    row["created_at"],
                    row["updated_at"],
                ]
            )

        movements_sheet.append(MOVEMENT_HEADERS)
        for row in data["movements"]:
            movements_sheet.append(
                [
                    row["id"],
                    row["item_id"],
                    row["created_at"],
                    excel_safe_text(row["item_name"]),
                    row["movement_type"],
                    raw_to_quantity(int(row["quantity"])),
                    raw_to_quantity(int(row["stock_before"])),
                    raw_to_quantity(int(row["stock_after"])),
                    excel_safe_text(row["note"]),
                ]
            )

        categories_sheet.append(CATEGORY_HEADERS)
        for row in data["category_records"]:
            categories_sheet.append(
                [
                    row["id"],
                    excel_safe_text(row["name"]),
                    row["created_at"],
                    row["updated_at"],
                ]
            )

        locations_sheet.append(LOCATION_HEADERS)
        for row in data["location_records"]:
            locations_sheet.append(
                [
                    row["id"],
                    excel_safe_text(row["name"]),
                    excel_safe_text(row["description"]),
                    row["created_at"],
                    row["updated_at"],
                ]
            )

        all_items = data["items"]
        global_minimum = self._global_minimum_raw()
        summary_rows = [
            ("Metric", "Value"),
            ("Total items", len(all_items)),
            (
                "Total stock",
                raw_to_quantity(sum(int(row["current_stock"]) for row in all_items)),
            ),
            (
                "Low-stock items",
                sum(
                    1
                    for row in all_items
                    if 0 < int(row["current_stock"]) <= global_minimum
                ),
            ),
            (
                "Out-of-stock items",
                sum(1 for row in all_items if int(row["current_stock"]) == 0),
            ),
            ("Backup creation time", created_at.isoformat().replace("+00:00", "Z")),
            (),
            ("Category totals",),
            ("Category", "Item Count", "Total Stock"),
        ]
        for values in summary_rows:
            summary_sheet.append(values)
        for row in data["categories"]:
            summary_sheet.append(
                [row["name"], int(row["item_count"]), raw_to_quantity(int(row["total_stock"]))]
            )
        summary_sheet.append(())
        summary_sheet.append(("Location totals",))
        summary_sheet.append(("Location", "Item Count", "Total Stock"))
        for row in data["locations"]:
            summary_sheet.append(
                [row["name"], int(row["item_count"]), raw_to_quantity(int(row["total_stock"]))]
            )

        metadata = [
            ("Key", "Value"),
            ("Application ID", "inventaris-gudang-local"),
            ("Application Version", __version__),
            ("Database Schema Version", self.database.schema_version()),
            ("Backup Format Version", 2),
            ("Backup Creation Time", created_at.isoformat().replace("+00:00", "Z")),
            ("Item Count", len(data["items"])),
            ("Movement Count", len(data["movements"])),
            ("Database File Name", self.config.database_path.name),
            ("Quantity Scale", 1000),
            ("Money Scale", MONEY_SCALE),
            ("Company Name", excel_safe_text(data["settings"].get("company_name", ""))),
        ]
        for row in metadata:
            metadata_sheet.append(row)

        settings_sheet.append(("Setting Key", "Value"))
        for key, value in sorted(data["settings"].items()):
            settings_sheet.append((excel_safe_text(key), excel_safe_text(value)))

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="234E70")
                cell.alignment = Alignment(vertical="center")
            for column in sheet.columns:
                letter = column[0].column_letter
                width = min(45, max(12, *(len(str(cell.value or "")) + 2 for cell in column)))
                sheet.column_dimensions[letter].width = width
        return workbook

    @staticmethod
    def _verify_workbook(path: Path) -> None:
        if not path.exists() or path.stat().st_size < 100:
            raise ValueError("Workbook file was not created correctly.")
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        try:
            required = {
                "Items",
                "Stock Movements",
                "Categories",
                "Locations",
                "Summary",
                "Backup Metadata",
                "Advanced Settings",
            }
            if not required.issubset(workbook.sheetnames):
                raise ValueError("Workbook is missing required sheets.")
            actual_headers = [cell.value for cell in next(workbook["Items"].iter_rows(max_row=1))]
            accepted = {
                tuple(ITEM_HEADERS),
                tuple(LEGACY_ITEM_HEADERS),
                tuple(LEGACY_ITEM_HEADERS_V1),
                tuple(LEGACY_ITEM_HEADERS_V2),
                tuple(LEGACY_ITEM_HEADERS_V3),
            }
            if tuple(actual_headers) not in accepted:
                raise ValueError("Items header verification failed.")
        finally:
            workbook.close()

    def _global_minimum_raw(self) -> int:
        from app.utils import QUANTITY_SCALE

        with self.database.connection() as connection:
            row = connection.execute(
                "SELECT value FROM app_settings WHERE key = ?",
                ("analytics.stock_risk.default_minimum",),
            ).fetchone()
        try:
            public = float(row["value"]) if row else 10
        except (TypeError, ValueError, KeyError):
            public = 10
        if public < 0:
            public = 0
        return int(round(public * QUANTITY_SCALE))

    def _start_log(self, path: Path) -> str:
        log_id = new_id()
        with self.database.transaction() as connection:
            connection.execute(
                """
                INSERT INTO backup_logs(
                    id, backup_type, file_name, file_path, status, error_message, created_at
                ) VALUES (?, 'EXCEL', ?, ?, 'RUNNING', NULL, ?)
                """,
                (
                    log_id,
                    path.name,
                    str(path.relative_to(self.config.root)),
                    utc_now(),
                ),
            )
        return log_id

    def _finish_log(self, log_id: str, status: str, error: str | None) -> None:
        with self.database.transaction() as connection:
            connection.execute(
                "UPDATE backup_logs SET status = ?, error_message = ? WHERE id = ?",
                (status, error, log_id),
            )

    def _cleanup_retention(self) -> None:
        with self.database.connection() as connection:
            row = connection.execute(
                "SELECT value FROM app_settings WHERE key = 'daily_backup_retention_days'"
            ).fetchone()
        days = int(row["value"]) if row else self.config.daily_backup_retention_days
        cutoff = datetime.now(UTC) - timedelta(days=max(1, days))
        for path in self.config.daily_backups_path.glob("inventory_*.xlsx"):
            try:
                modified = datetime.fromtimestamp(path.stat().st_mtime, UTC)
                if modified < cutoff:
                    path.unlink()
            except OSError:
                logger.warning("Could not delete expired backup %s", path.name)
