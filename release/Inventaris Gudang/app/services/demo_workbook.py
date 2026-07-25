"""Portable workbook writer for the deterministic demo dataset."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app import __version__
from app.services.demo_dataset import DEMO_SEED, DEMO_VERSION, DemoDataset
from app.utils import MONEY_SCALE, QUANTITY_SCALE, excel_safe_text


def write_demo_workbook(dataset: DemoDataset, destination: Path) -> None:
    """Create a professionally formatted, macro-free workbook."""

    from app.services.analytics_config import (
        CHART_SETTING_KEYS,
        DEFAULT_SETTINGS,
        SETTING_KEYS,
    )
    from app.services.backup_excel import (
        CATEGORY_HEADERS,
        ITEM_HEADERS,
        LOCATION_HEADERS,
        MOVEMENT_HEADERS,
    )

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    items_sheet = workbook.create_sheet("Items")
    movement_sheet = workbook.create_sheet("Stock Movements")
    category_sheet = workbook.create_sheet("Categories")
    location_sheet = workbook.create_sheet("Locations")
    summary_sheet = workbook.create_sheet("Summary")
    metadata_sheet = workbook.create_sheet("Backup Metadata")
    settings_sheet = workbook.create_sheet("Advanced Settings")

    readme_rows = [
        ("Agricultural Inventory Demonstration Dataset",),
        ("Status", "FICTIONAL DEMONSTRATION DATA"),
        ("Purpose", "Offline analytics and workflow demonstration."),
        ("Dataset Version", DEMO_VERSION),
        ("Generator Seed", DEMO_SEED),
        ("Generated At", dataset.generated_at),
        ("Important", "No real brands, customers, or transactions are represented."),
    ]
    for row in readme_rows:
        readme.append(row)

    items_sheet.append(ITEM_HEADERS)
    for row in dataset.items:
        items_sheet.append(
            [
                row["id"],
                excel_safe_text(row["name"]),
                excel_safe_text(row["category_name"]),
                excel_safe_text(row["location_name"]),
                row["unit"],
                row["current_stock"] / QUANTITY_SCALE,
                excel_safe_text(row["description"]),
                row["created_at"],
                row["updated_at"],
            ]
        )

    movement_sheet.append(MOVEMENT_HEADERS)
    item_by_id = {row["id"]: row for row in dataset.items}
    for row in dataset.movements:
        item = item_by_id[row["item_id"]]
        movement_sheet.append(
            [
                row["id"],
                row["item_id"],
                row["created_at"],
                item["name"],
                row["movement_type"],
                row["quantity"] / QUANTITY_SCALE,
                row["stock_before"] / QUANTITY_SCALE,
                row["stock_after"] / QUANTITY_SCALE,
                excel_safe_text(row["note"]),
            ]
        )

    category_sheet.append(CATEGORY_HEADERS)
    for row in dataset.categories:
        category_sheet.append(
            [row["id"], row["name"], row["created_at"], row["updated_at"]]
        )
    location_sheet.append(LOCATION_HEADERS)
    for row in dataset.locations:
        location_sheet.append(
            [
                row["id"],
                row["name"],
                row["description"],
                row["created_at"],
                row["updated_at"],
            ]
        )
    summary_sheet.append(("Metric", "Value"))
    for key, value in dataset.counts.items():
        summary_sheet.append((key.replace("_", " ").title(), value))
    summary_sheet.append(("History Start", min(row["created_at"] for row in dataset.movements)))
    summary_sheet.append(("History End", max(row["created_at"] for row in dataset.movements)))
    summary_sheet.append(("Generator Seed", DEMO_SEED))

    metadata = (
        ("Key", "Value"),
        ("Application ID", "inventaris-gudang-local"),
        ("Application Version", __version__),
        ("Database Schema Version", 5),
        ("Backup Format Version", 2),
        ("Backup Creation Time", dataset.generated_at),
        ("Item Count", len(dataset.items)),
        ("Movement Count", len(dataset.movements)),
        ("Database File Name", "inventory.db"),
        ("Quantity Scale", QUANTITY_SCALE),
        ("Money Scale", MONEY_SCALE),
        ("Dataset Version", DEMO_VERSION),
        ("Generator Seed", DEMO_SEED),
        ("Demonstration Data", "true"),
    )
    for row in metadata:
        metadata_sheet.append(row)

    settings_sheet.append(("Setting Key", "Value"))

    def setting_value(value: Any) -> str:
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, (list, dict)):
            return json.dumps(value, ensure_ascii=False)
        return str(value)

    settings_rows = {
        setting_key: setting_value(DEFAULT_SETTINGS[name])
        for name, setting_key in SETTING_KEYS.items()
    }
    settings_rows.update(
        {
            setting_key: setting_value(DEFAULT_SETTINGS["chart_visibility"][chart_id])
            for chart_id, setting_key in CHART_SETTING_KEYS.items()
        }
    )
    settings_rows.update(
        {
            "company_name": "ALFAN TANI",
            "demo.auto_load_disabled": "false",
            "demo.dataset_version": DEMO_VERSION,
        }
    )
    for key, value in sorted(settings_rows.items()):
        settings_sheet.append((key, value))

    currency_headers = {"Purchase Price", "Selling Price"}
    quantity_headers = {
        "Current Stock",
        "Minimum Stock",
        "Quantity",
        "Stock Before",
        "Stock After",
    }
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row > 1 and sheet.max_column > 1:
            sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="245C72")
            cell.alignment = Alignment(vertical="center")
        for column in sheet.columns:
            cells = list(column)
            letter = cells[0].column_letter
            width = min(48, max(12, *(len(str(cell.value or "")) + 2 for cell in cells)))
            sheet.column_dimensions[letter].width = width
        headers = {cell.column: str(cell.value or "") for cell in sheet[1]}
        for column_index, header in headers.items():
            if header in currency_headers:
                for cell in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                ):
                    for value_cell in cell:
                        value_cell.number_format = '"Rp" #,##0'
            elif header in quantity_headers:
                for cell in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                ):
                    for value_cell in cell:
                        value_cell.number_format = "#,##0.###"
            elif "Date" in header or "Time" in header:
                for cell in sheet.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                ):
                    for value_cell in cell:
                        value_cell.number_format = "yyyy-mm-dd"

    temporary = destination.with_suffix(".tmp.xlsx")
    workbook.save(temporary)
    workbook.close()
    temporary.replace(destination)
