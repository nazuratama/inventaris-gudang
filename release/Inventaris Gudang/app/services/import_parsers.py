"""Excel, CSV, and legacy JSON import parsers."""

from __future__ import annotations

import csv
import io
import json
import zipfile
from datetime import datetime
from pathlib import PurePosixPath
from typing import Any

from openpyxl import load_workbook

from app.services.backup_excel import (
    CATEGORY_HEADERS,
    ITEM_HEADERS,
    LEGACY_CATEGORY_HEADERS,
    LEGACY_ITEM_HEADERS,
    LEGACY_ITEM_HEADERS_V1,
    LEGACY_ITEM_HEADERS_V2,
    LEGACY_ITEM_HEADERS_V3,
    LEGACY_LOCATION_HEADERS,
    LEGACY_MOVEMENT_HEADERS,
    LEGACY_MOVEMENT_HEADERS_V1,
    LEGACY_MOVEMENT_HEADERS_V2,
    LOCATION_HEADERS,
    MOVEMENT_HEADERS,
)
from app.errors import AppError
from app.utils import new_id, normalize_text, utc_now


class ImportParsers:
    def _parse_excel(self, content: bytes) -> dict[str, Any]:
        self._validate_xlsx_container(content)
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        try:
            required = {"Items", "Stock Movements", "Summary", "Backup Metadata"}
            if not required.issubset(workbook.sheetnames):
                missing = sorted(required.difference(workbook.sheetnames))
                raise AppError(
                    "MISSING_WORKBOOK_SHEETS",
                    "Workbook tidak memiliki seluruh sheet yang diperlukan.",
                    status_code=422,
                    details={"missing": missing},
                )
            metadata_sheet = workbook["Backup Metadata"]
            metadata_rows = list(metadata_sheet.iter_rows())
            self._reject_formula_cells(metadata_rows)
            metadata = {
                str(row[0].value): row[1].value
                for row in metadata_rows[1:]
                if len(row) >= 2 and row[0].value
            }
            self._validate_backup_metadata(metadata)
            format_version = int(metadata["Backup Format Version"])
            if format_version >= 2:
                extended = {
                    "Categories",
                    "Locations",
                    "Advanced Settings",
                }
                missing_extended = sorted(extended.difference(workbook.sheetnames))
                if missing_extended:
                    raise AppError(
                        "MISSING_WORKBOOK_SHEETS",
                        "Workbook versi baru tidak memiliki seluruh sheet yang diperlukan.",
                        status_code=422,
                        details={"missing": missing_extended},
                    )

            item_sheet = workbook["Items"]
            movement_sheet = workbook["Stock Movements"]
            item_rows = list(item_sheet.iter_rows())
            movement_rows = list(movement_sheet.iter_rows())
            item_headers = [cell.value for cell in item_rows[0]] if item_rows else []
            movement_headers = [cell.value for cell in movement_rows[0]] if movement_rows else []
            accepted_item_headers = [
                ITEM_HEADERS,
                LEGACY_ITEM_HEADERS,
                LEGACY_ITEM_HEADERS_V1,
                LEGACY_ITEM_HEADERS_V2,
                LEGACY_ITEM_HEADERS_V3,
            ]
            accepted_movement_headers = [
                MOVEMENT_HEADERS,
                LEGACY_MOVEMENT_HEADERS,
                LEGACY_MOVEMENT_HEADERS_V1,
                LEGACY_MOVEMENT_HEADERS_V2,
            ]
            if item_headers not in accepted_item_headers:
                raise AppError(
                    "INVALID_ITEMS_SHEET",
                    "Struktur sheet Items tidak didukung.",
                    status_code=422,
                )
            if list(movement_headers) not in accepted_movement_headers:
                raise AppError(
                    "INVALID_MOVEMENTS_SHEET",
                    "Struktur sheet Stock Movements tidak didukung.",
                    status_code=422,
                )
            self._reject_formula_cells(item_rows[1:])
            self._reject_formula_cells(movement_rows[1:])
            simple_items = item_headers == ITEM_HEADERS
            legacy_simple_active = item_headers == LEGACY_ITEM_HEADERS_V3
            legacy_with_type = item_headers == LEGACY_ITEM_HEADERS_V2
            legacy_priced = item_headers == LEGACY_ITEM_HEADERS
            items: list[dict[str, Any]] = []
            for row_number, cells in enumerate(item_rows[1:], start=2):
                values = [self._restore_excel_text(cell.value) for cell in cells]
                if all(value in (None, "") for value in values):
                    continue
                if simple_items:
                    # ID, Name, Category, Location, Unit, Stock, Description, Created, Updated
                    items.append(
                        {
                            "source_row": row_number,
                            "id": values[0],
                            "name": values[1],
                            "category": values[2],
                            "location": values[3],
                            "unit": values[4],
                            "current_stock": values[5],
                            "description": values[6],
                            "created_at": values[7],
                            "updated_at": values[8],
                        }
                    )
                elif legacy_simple_active:
                    # Previous simple export with Active Status (ignored).
                    items.append(
                        {
                            "source_row": row_number,
                            "id": values[0],
                            "name": values[1],
                            "category": values[2],
                            "location": values[3],
                            "unit": values[4],
                            "current_stock": values[5],
                            "description": values[6],
                            "created_at": values[8],
                            "updated_at": values[9],
                        }
                    )
                elif legacy_with_type:
                    items.append(
                        {
                            "source_row": row_number,
                            "id": values[0],
                            "name": values[1],
                            "category": values[2],
                            "location": values[3],
                            "unit": values[4],
                            "current_stock": values[5],
                            "description": values[6],
                            "created_at": values[9],
                            "updated_at": values[10],
                        }
                    )
                elif legacy_priced:
                    items.append(
                        {
                            "source_row": row_number,
                            "id": values[0],
                            "name": values[2],
                            "category": values[3],
                            "location": values[4],
                            "unit": values[5],
                            "current_stock": values[6],
                            "description": values[12],
                            "created_at": values[15],
                            "updated_at": values[16],
                        }
                    )
                else:
                    # LEGACY_ITEM_HEADERS_V1
                    items.append(
                        {
                            "source_row": row_number,
                            "id": values[0],
                            "name": values[2],
                            "category": values[3],
                            "location": values[4],
                            "unit": values[5],
                            "current_stock": values[6],
                            "description": values[9],
                            "created_at": values[11],
                            "updated_at": values[12],
                        }
                    )
            movements: list[dict[str, Any]] = []
            movement_header_list = list(movement_headers)
            simple_movements = movement_header_list == MOVEMENT_HEADERS
            for row_number, cells in enumerate(movement_rows[1:], start=2):
                values = [self._restore_excel_text(cell.value) for cell in cells]
                if all(value in (None, "") for value in values):
                    continue
                if simple_movements:
                    movements.append(
                        {
                            "source_row": row_number,
                            "id": values[0],
                            "item_id": values[1],
                            "created_at": values[2],
                            "movement_type": values[4],
                            "quantity": values[5],
                            "stock_before": values[6],
                            "stock_after": values[7],
                            "note": values[8],
                            "reference_number": None,
                        }
                    )
                    continue
                if movement_header_list == LEGACY_MOVEMENT_HEADERS_V2:
                    movements.append(
                        {
                            "source_row": row_number,
                            "id": values[0],
                            "item_id": values[1],
                            "created_at": values[2],
                            "movement_type": values[4],
                            "quantity": values[5],
                            "stock_before": values[6],
                            "stock_after": values[7],
                            "note": values[8],
                            "reference_number": None,
                        }
                    )
                    continue
                movements.append(
                    {
                        "source_row": row_number,
                        "id": values[0],
                        "item_id": values[1],
                        "created_at": values[2],
                        "movement_type": values[5],
                        "quantity": values[6],
                        "stock_before": values[7],
                        "stock_after": values[8],
                        "note": values[9],
                        "reference_number": values[10],
                    }
                )
            categories = (
                self._parse_named_sheet(
                    workbook["Categories"],
                    CATEGORY_HEADERS,
                    LEGACY_CATEGORY_HEADERS,
                )
                if format_version >= 2
                else []
            )
            locations = (
                self._parse_named_sheet(
                    workbook["Locations"],
                    LOCATION_HEADERS,
                    LEGACY_LOCATION_HEADERS,
                )
                if format_version >= 2
                else []
            )
            settings = {}
            if format_version >= 2:
                setting_rows = list(workbook["Advanced Settings"].iter_rows())
                self._reject_formula_cells(setting_rows)
                settings = {
                    str(row[0].value): str(row[1].value)
                    for row in setting_rows[1:]
                    if len(row) >= 2 and row[0].value is not None
                }
        finally:
            workbook.close()
        return {
            "format": "APPLICATION_XLSX",
            "mode": "RESTORE",
            "items": items,
            "movements": movements,
            "categories": categories,
            "locations": locations,
            "settings": settings,
            "warnings": [],
        }

    def _parse_csv(self, content: bytes) -> dict[str, Any]:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise AppError(
                "INVALID_CSV_ENCODING",
                "CSV harus menggunakan encoding UTF-8.",
                status_code=422,
            ) from exc
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise AppError("INVALID_CSV", "CSV tidak memiliki header.", status_code=422)
        aliases = {
            "sku": ("sku",),
            "name": ("name", "nama", "nama barang"),
            "category": ("category", "kategori"),
            "location": ("location", "lokasi"),
            "unit": ("unit", "satuan"),
            "current_stock": ("current stock", "stok", "stok saat ini"),
            "minimum_stock": ("minimum stock", "stok minimum"),
            "description": ("description", "deskripsi"),
        }
        header_map = {str(header).strip().casefold(): header for header in reader.fieldnames}
        mapped: dict[str, str | None] = {}
        for target, choices in aliases.items():
            mapped[target] = next((header_map[key] for key in choices if key in header_map), None)
        if not mapped["name"] or not mapped["current_stock"]:
            raise AppError(
                "MISSING_CSV_COLUMNS",
                "CSV wajib memiliki kolom nama barang dan stok.",
                status_code=422,
            )
        items = []
        for row_number, row in enumerate(reader, start=2):
            values = {
                key: row.get(header, "") if header else None for key, header in mapped.items()
            }
            if not normalize_text(values.get("unit"), empty_to_none=True):
                values["unit"] = "Pcs"
            if values.get("minimum_stock") in (None, ""):
                values["minimum_stock"] = 0
            items.append(
                {
                    "source_row": row_number,
                    "id": new_id(),
                    **values,
                    "is_active": True,
                    "created_at": utc_now(),
                    "updated_at": utc_now(),
                }
            )
        return {
            "format": "CSV",
            "mode": "IMPORT",
            "items": items,
            "movements": [],
            "warnings": [],
        }

    def _parse_legacy_json(self, content: bytes) -> dict[str, Any]:
        try:
            decoded = json.loads(content.decode("utf-8-sig"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise AppError(
                "INVALID_JSON",
                "File JSON tidak valid.",
                status_code=422,
            ) from exc
        if isinstance(decoded, dict):
            if decoded.get("schema_version") != "legacy-inventory-1":
                raise AppError(
                    "UNSUPPORTED_LEGACY_SCHEMA",
                    "Versi skema JSON legacy tidak didukung.",
                    status_code=422,
                )
            source = decoded.get("items")
        else:
            # Raw arrays are accepted only for direct exports from the original prototype.
            source = decoded
        if not isinstance(source, list):
            raise AppError(
                "INVALID_LEGACY_SCHEMA",
                "Struktur JSON legacy tidak didukung.",
                status_code=422,
            )
        items = []
        for row_number, row in enumerate(source, start=1):
            if not isinstance(row, dict):
                items.append({"source_row": row_number, "_invalid": True})
                continue
            items.append(
                {
                    "source_row": row_number,
                    "id": new_id(),
                    "sku": row.get("sku"),
                    "name": row.get("nama", row.get("name")),
                    "category": row.get("kategori", row.get("category")),
                    "location": row.get("lokasi", row.get("location")),
                    "unit": row.get("satuan", row.get("unit", "Pcs")),
                    "current_stock": row.get("stok", row.get("current_stock", 0)),
                    "minimum_stock": row.get("minimum_stock", 0),
                    "description": row.get("deskripsi", row.get("description")),
                    "is_active": True,
                    "created_at": utc_now(),
                    "updated_at": utc_now(),
                }
            )
        return {
            "format": "LEGACY_JSON",
            "mode": "IMPORT",
            "items": items,
            "movements": [],
            "warnings": [],
        }

    def _parse_named_sheet(
        self,
        sheet: Any,
        expected_headers: list[str],
        *legacy_headers: list[str],
    ) -> list[dict[str, Any]]:
        rows = list(sheet.iter_rows())
        actual = [cell.value for cell in rows[0]] if rows else []
        accepted = [expected_headers, *legacy_headers]
        if not rows or actual not in accepted:
            raise AppError(
                "INVALID_WORKBOOK_SHEET",
                f"Struktur sheet {sheet.title} tidak didukung.",
                status_code=422,
            )
        self._reject_formula_cells(rows[1:])
        result = []
        for row_number, cells in enumerate(rows[1:], start=2):
            values = [self._restore_excel_text(cell.value) for cell in cells]
            if all(value in (None, "") for value in values):
                continue
            result.append(
                {
                    "source_row": row_number,
                    "values": values,
                    "headers": actual,
                }
            )
        return result

    def _validate_xlsx_container(self, content: bytes) -> None:
        stream = io.BytesIO(content)
        if not zipfile.is_zipfile(stream):
            raise AppError("INVALID_XLSX", "File XLSX tidak valid.", status_code=422)
        stream.seek(0)
        with zipfile.ZipFile(stream) as archive:
            entries = archive.infolist()
            if len(entries) > 500:
                raise AppError("XLSX_TOO_COMPLEX", "Workbook terlalu kompleks.", status_code=422)
            total = 0
            for entry in entries:
                path = PurePosixPath(entry.filename)
                if path.is_absolute() or ".." in path.parts:
                    raise AppError(
                        "INVALID_XLSX_PATH",
                        "Struktur workbook tidak aman.",
                        status_code=422,
                    )
                total += entry.file_size
                if total > self.config.maximum_excel_uncompressed_bytes:
                    raise AppError(
                        "XLSX_UNCOMPRESSED_TOO_LARGE",
                        "Ukuran workbook setelah dibuka melebihi batas.",
                        status_code=422,
                    )
                if entry.compress_size > 0 and entry.file_size / entry.compress_size > 1000:
                    raise AppError(
                        "XLSX_COMPRESSION_RATIO_INVALID",
                        "Workbook memiliki rasio kompresi yang tidak aman.",
                        status_code=422,
                    )

    def _validate_backup_metadata(self, metadata: dict[str, Any]) -> None:
        required = {
            "Application ID",
            "Application Version",
            "Database Schema Version",
            "Backup Format Version",
            "Backup Creation Time",
            "Quantity Scale",
        }
        missing = sorted(required.difference(metadata))
        if missing:
            raise AppError(
                "INVALID_BACKUP_METADATA",
                "Metadata backup tidak lengkap.",
                status_code=422,
                details={"missing": missing},
            )
        try:
            schema_version = int(metadata["Database Schema Version"])
            format_version = int(metadata["Backup Format Version"])
            quantity_scale = int(metadata["Quantity Scale"])
        except (TypeError, ValueError) as exc:
            raise AppError(
                "INVALID_BACKUP_METADATA",
                "Versi atau skala pada metadata backup tidak valid.",
                status_code=422,
            ) from exc
        if metadata["Application ID"] != "inventaris-gudang-local":
            raise AppError(
                "FOREIGN_BACKUP",
                "File bukan backup Inventaris Gudang yang didukung.",
                status_code=422,
            )
        if format_version not in {1, 2} or quantity_scale != 1000:
            raise AppError(
                "UNSUPPORTED_BACKUP_FORMAT",
                "Versi format backup tidak didukung.",
                status_code=422,
            )
        if format_version == 2:
            try:
                money_scale = int(metadata["Money Scale"])
            except (KeyError, TypeError, ValueError) as exc:
                raise AppError(
                    "INVALID_BACKUP_METADATA",
                    "Skala harga pada metadata backup tidak valid.",
                    status_code=422,
                ) from exc
            if money_scale != 100:
                raise AppError(
                    "UNSUPPORTED_BACKUP_FORMAT",
                    "Skala harga backup tidak didukung.",
                    status_code=422,
                )
        if schema_version < 1 or schema_version > self.database.schema_version():
            raise AppError(
                "UNSUPPORTED_SCHEMA_VERSION",
                "Versi skema database pada backup tidak didukung.",
                status_code=422,
            )
        if not self._valid_iso_timestamp(metadata["Backup Creation Time"]):
            raise AppError(
                "INVALID_BACKUP_TIMESTAMP",
                "Waktu pembuatan backup tidak valid.",
                status_code=422,
            )

    @staticmethod
    def _reject_formula_cells(rows: list[tuple[Any, ...]]) -> None:
        for row in rows:
            if any(cell.data_type == "f" for cell in row):
                raise AppError(
                    "FORMULA_NOT_ALLOWED",
                    "Workbook impor tidak boleh berisi formula.",
                    status_code=422,
                )

    @staticmethod
    def _restore_excel_text(value: Any) -> Any:
        if isinstance(value, str) and len(value) >= 2 and value[0] == "'" and value[1] in "=+-@":
            return value[1:]
        return value
