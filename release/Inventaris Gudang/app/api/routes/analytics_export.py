"""Analytics tabular export endpoint."""

from __future__ import annotations

import csv
import io
from datetime import date
from typing import Any, Literal

from fastapi import APIRouter, Query, Request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from starlette.responses import StreamingResponse

from app.api.analytics_dependencies import _filters, _service
from app.utils import excel_safe_text
from app.validation.analytics import ChartId

router = APIRouter(prefix="/api/v1/analytics")


@router.get("/export")
async def export_chart(
    request: Request,
    chart_id: ChartId,
    format: Literal["csv", "xlsx"] = "csv",
    date_range: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    category_id: str | None = None,
    location_id: str | None = None,
    include_archived: bool | None = None,
    include_demo: bool | None = None,
    data_scope: Literal["all", "demo", "real"] = "all",
    aggregation: Literal["daily", "weekly", "monthly"] | None = None,
    top_n: int | None = Query(default=None, ge=5, le=20),
    ranking: Literal["highest", "lowest"] = "highest",
    movement_scope: Literal["both", "in", "out"] = "both",
    metric: str | None = Query(default=None, max_length=30),
    show_net: bool = False,
) -> StreamingResponse:
    filters = _filters(
        request,
        date_range=date_range,
        date_from=date_from,
        date_to=date_to,
        category_id=category_id,
        location_id=location_id,
        include_archived=include_archived,
        include_demo=include_demo,
        data_scope=data_scope,
        aggregation=aggregation,
        top_n=top_n,
        ranking=ranking,
        movement_scope=movement_scope,
        metric=metric,
        show_net=show_net,
    )
    data = _service(request).chart(chart_id, filters)
    rows = data["table_rows"]
    headers = list(rows[0]) if rows else ["message"]
    if format == "csv":
        stream = io.StringIO()
        writer = csv.DictWriter(stream, fieldnames=headers)
        writer.writeheader()
        if rows:
            for row in rows:
                writer.writerow({key: excel_safe_text(row.get(key)) for key in headers})
        else:
            writer.writerow({"message": "Tidak ada data"})
        content = io.BytesIO(stream.getvalue().encode("utf-8-sig"))
        media_type = "text/csv; charset=utf-8"
        suffix = "csv"
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Chart Data"
        sheet.append(headers)
        for row in rows:
            sheet.append([excel_safe_text(row.get(key)) for key in headers])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="245C72")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        content = io.BytesIO()
        workbook.save(content)
        workbook.close()
        content.seek(0)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        suffix = "xlsx"
    filename = f"{chart_id}_{date.today().isoformat()}.{suffix}"
    return StreamingResponse(
        content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
